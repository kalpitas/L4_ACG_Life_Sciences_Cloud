from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

wb = load_workbook("ACG_Common_Workbook.xlsx")
typeev = wb["URL_Login_cred_Tenant"]
type_env = typeev.cell(row = 2, column = 4).value


driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(60)


if type_env == "ACG":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_ACG"]

if type_env == "Tenant":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_Tenant"]

url = url_login_cred.cell(row = 2, column = 1).value
print(url)
driver.get(str(url))


#if type_env == 'ACG':
driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()


usern = url_login_cred.cell(row = 2, column = 2).value
#if type_env == 'ACG':
driver.find_element(By.NAME,'userName').send_keys(str(usern))

#if type_env == 'Tenant':
#    driver.find_element(By.NAME,'email').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[2]/div/button").click()

roles_sheet = wb['Rights_test_case']
start_role = 8
last_role = 10
#(role.max_row) +1

if type_env == 'ACG':
    
    driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[1]/div/div/button[3]").click()
    driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[1]/div/div/button[3]/div").click()

if type_env == 'Tenant':
    
    driver.find_element(By.XPATH,"/html/body/div[1]/div/div/div[2]/div[1]/div/div[4]/h2/button").click()
    driver.find_element(By.XPATH,"/html/body/div[1]/div/div/div[2]/div[1]/div/div[4]/div/div/a[2]").click()


for i in range(start_role, last_role):

    timeout = 10
    try:
        new_role = EC.element_to_be_clickable((By.XPATH,"/html/body/div[1]/div/div/div[2]/div[2]/div/div/div[2]/b"))
        WebDriverWait(driver,timeout).until(new_role)
    except TimeoutException:
        print("Create role page: Timed out waiting for page to load")
    
    time.sleep(5)
    
    driver.find_element(By.XPATH,"/html/body/div[1]/div/div/div[2]/div[2]/div/div/div[2]/b").click()
    
    
    timeout = 10
    try:
        new_user_nxt = EC.element_to_be_clickable((By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]"))
        WebDriverWait(driver,timeout).until(new_user_nxt)
    except TimeoutException:
        print("Create role page_submit: Timed out waiting for page to load")
    
    
    r_name = roles_sheet.cell(row = i, column = 3).value
    print(r_name)
    driver.find_element(By.NAME,"roleName").send_keys(str(r_name))
    
    role_desp = roles_sheet.cell(row = i, column = 4).value
    driver.find_element(By.NAME,"roleDescription").send_keys(str(role_desp))
    

    roles = driver.find_elements(By.XPATH,"//span[starts-with(@class,'chakra-checkbox__label')]")
    #//input[contains(@class,"chakra-checkbox__input")]
    if i == 8:
        
        for checkbox in roles:
            role = checkbox.text
            print(role)
            
            Role_management = {"Create New Role":1,"Get Roles with Privileges":2,"Get All Roles for Tenant":3,
                        "Update Roles":4}
            
            User_management = {"Get User List":1,"Update User Details":2,
                        "Create New Tenant User":3,"View Single User Details":4}
            
            EPCIS = {"Get All Events Based on EPC":1,"Get All Events Based on Location Identifier":2,
                        "Get All Events Tenant":3,"Get EPICS File Tenant":4,
                        "Get Single Event Epcis2":5,"Get Events by ID":6,"Upload Files to Data Lake":7}
            
            Master_M = {"Create Product":1,"Get All Product List Details":2,
                        "Get All Products":3,"Get Product Details":4,
                        "Update Location":5,"Get All Locations for Tenant":6,
                        "Create New Location":7,"View Single Product by ID":8,
                        "Get Location by ID":9,"Update Product":10}
            
            Integration = {"Configure SFTP User Tenant":1,
                        "Get All Integrations":2,"View Single Integration":3}
            
            Audit = {"Retrieve All Audit Files":1,"Get Audit Logs By Time":2}
            
            if role in Role_management:
                labell = 4
                ide = Role_management[f'{role}']
   
            if role in User_management:
                labell = 5
                ide = User_management[f'{role}']
                
            if role in EPCIS:
                labell = 6
                ide = EPCIS[f'{role}']
            
            if role in Master_M:
                labell = 7
                ide = Master_M[f'{role}']
                
   
            if role in Integration:
                labell = 8
                ide = Integration[f'{role}']
                
            if role in Audit:
                labell = 9
                ide = Audit[f'{role}']
            
            print(ide)
            checkbox1 = driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div/div["+str(labell)+"]/div/label["+str(ide)+"]/span[2]")
            driver.execute_script("arguments[0].scrollIntoView()", checkbox1)
            driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div/div["+str(labell)+"]/div/label["+str(ide)+"]/span[2]").click()

    if i == 9:
        
        for checkbox in roles:
            role = checkbox.text
            print(role)
            
            Role_management = {"Create New Role":1,"Get Roles with Privileges":2,"Get All Roles for Tenant":3,
                        }
            
            User_management = {"Get User List":1,
                        "Create New Tenant User":3,"View Single User Details":4}
            
            Master_M = {"Create Product":1,"Get All Product List Details":2,
                        "Get All Products":3,"Get Product Details":4,
                        "Get All Locations for Tenant":6,
                        "Create New Location":7,"View Single Product by ID":8,
                        "Get Location by ID":9}
            
            Audit = {"Retrieve All Audit Files":1,"Get Audit Logs By Time":2}
            
            if role in Role_management:
                labell = 4
                ide = Role_management[f'{role}']
   
            if role in User_management:
                labell = 5
                ide = User_management[f'{role}']
            
            if role in Master_M:
                labell = 7
                ide = Master_M[f'{role}']
                
            if role in Audit:
                labell = 9
                ide = Audit[f'{role}']
            
            print(ide)
            checkbox1 = driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div/div["+str(labell)+"]/div/label["+str(ide)+"]/span[2]")
            driver.execute_script("arguments[0].scrollIntoView()", checkbox1)
            driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div/div["+str(labell)+"]/div/label["+str(ide)+"]/span[2]").click()

            
    driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[1]/button[2]").click()
    timeout = 2
    try:
        new_user = EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div/div[2]/div[2]/div/div/div[2]/b"))
        WebDriverWait(driver, timeout).until(new_user)
        message = driver.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/div/div/div[2]").text
    except TimeoutException:
        driver.find_element(By.XPATH, "//*[@id='root']/div/div/div[2]/div[2]/div/div[1]/button[2]").click()
        time.sleep(2)
        message = driver.find_element(By.XPATH,"/html/body/div[2]/div[4]/div/div/div/div/div[2]").text
        driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[1]/button[1]").click()

    roles_sheet.cell(row = i, column = 5).value = message
    wb.save("ACG_Common_Workbook.xlsx")