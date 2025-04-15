from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

    
driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(60)

        
wb = load_workbook("ACG_Common_Workbook.xlsx")
typeev = wb["URL_Login_cred_Tenant"]
type_env = typeev.cell(row = 2, column = 4).value


if type_env == "ACG":

    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_ACG"]
    start_row_log = 2
    last_row_log = (url_login_cred.max_row) + 1

if type_env == "Tenant":

    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_Tenant"]
    start_row_log = 2
    last_row_log = (url_login_cred.max_row) + 1
    


for i in range(start_row_log, last_row_log):
    
    
    url = url_login_cred.cell(row = i, column = 1).value
    print(url)
    driver.get(str(url))
    
    if type_env == 'ACG':
        driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()
    
    usern = url_login_cred.cell(row = i, column = 2).value
    driver.find_element(By.NAME,'userName').send_keys(str(usern))
    
    passw = url_login_cred.cell(row = i, column = 3).value
    driver.find_element(By.NAME,'password').send_keys(str(passw))
    
    driver.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[2]/div/button").click()
    
    time.sleep(5)
    print("done")
    print("to pull data")
    print("ti test branches")

