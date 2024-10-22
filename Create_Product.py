import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By


def main():

    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.implicitly_wait(60)

    workbook = load_workbook("ACG_Common_Workbook.xlsx")
    sheet = workbook["Product"]
    url_login_cred = workbook["URL_Login_cred_Tenant"]

    start_row = 2

    time.sleep(3)
    driver.get("https://proud-mud-040601f00.4.azurestaticapps.net/")

    driver.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

    usern = url_login_cred.cell(row=2, column=2).value
    # if type_env == 'ACG':
    driver.find_element(By.NAME, 'userName').send_keys(str(usern))

    passw = url_login_cred.cell(row=2, column=3).value
    driver.find_element(By.NAME, 'password').send_keys(str(passw))

    login = driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[4]/div/div[2]/div/button")
    login.click()
    time.sleep(6)

    # Navigate to the Add Product page
    EC.visibility_of_element_located((By.XPATH, "//span[text()= 'Master Data']"))
    driver.find_element(By.XPATH, "//span[text()= 'Master Data']").click()

    driver.find_element(By.XPATH,"//a[@href='/get/products']").click()

    # Loop through all the rows in the Excel sheet

    print(sheet.max_row+1)
    for i in range(start_row, sheet.max_row + 1):
        # Click 'Add Product' button at the beginning of each loop
        time.sleep(3)

        driver.find_element(By.XPATH,"/html/body/div[1]/div/div/div[2]/div[2]/div/div/div[2]/b").click()

        
        ProductIdentifier = sheet.cell(row = i, column = 1).value
        productName = sheet.cell(row = i, column = 2).value
        productDescription = sheet.cell(row = i, column = 3).value
        manufacturerName = sheet.cell(row = i, column = 4).value
        GS1CompanyPrefix = sheet.cell(row = i, column = 5).value
        GLN = sheet.cell(row = i, column = 6).value
        ProductIdentifier2 = sheet.cell(row = i, column = 7).value
        packagingLevelIndicator = sheet.cell(row = i, column = 8).value
        genericName = sheet.cell(row = i, column = 9).value
        MinTemp = sheet.cell(row = i, column = 10).value
        weight = sheet.cell(row = i, column = 11).value
        strength = sheet.cell(row = i, column = 12).value

        productIdentifierType = driver.find_element(By.XPATH, "//select[@name= 'productIdentifierType']")
        Select(productIdentifierType).select_by_index(2)

        # Fill in the Product details
        driver.find_element(By.XPATH, "//input[@name='productIdentifier']").send_keys(ProductIdentifier)

        driver.find_element(By.XPATH, "//input[@name='productName']").send_keys(productName)

        driver.find_element(By.XPATH, "//input[@name='productDescription']").send_keys(productDescription)


        checkbox1 = driver.find_element(By.XPATH,"//input[@name='productName']")
        driver.execute_script("arguments[0].scrollIntoView()", checkbox1)

        manufacturer = driver.find_element(By.XPATH, "//select[@name= 'manufacturer']")
        manufacturer.click()
        Select(manufacturer).select_by_visible_text("Others")

        driver.find_element(By.XPATH, "//input[@name='manufacturerName']").send_keys(manufacturerName)


        driver.find_element(By.XPATH, "//input[@name='manufacturerOtherGCP']").send_keys(GS1CompanyPrefix)


        driver.find_element(By.XPATH, "//b[text()='+ Add GLN ']").click()
        driver.find_element(By.XPATH, "//input[@placeholder='Enter GLN']").send_keys(GLN)


        # Next to Packaging Details
        driver.find_element(By.XPATH, "//button[text()= 'Next']").click()


        packagingCodeType = driver.find_element(By.XPATH, "//select[@name='packagingCodeType']")
        Select(packagingCodeType).select_by_index(1)


        driver.find_element(By.XPATH, "//input[@name='packagingProductIdentifier']").send_keys(ProductIdentifier2)

        Select(driver.find_element(By.XPATH, "//select[@name='packagingLevel']")).select_by_index(1)

        driver.find_element(By.XPATH, "//input[@name='packagingLevelIndicator']").send_keys(packagingLevelIndicator)


        driver.find_element(By.XPATH, "//button[text()= 'Add']").click()


        # Next to Other Details
        driver.find_element(By.XPATH, "//button[text()= 'Next']").click()


        driver.find_element(By.XPATH, "//input[@placeholder='Enter generic name']").send_keys(genericName)

        driver.find_element(By.XPATH, "//input[@placeholder='Enter min temperature']").send_keys(MinTemp)


        # Next to Regulatory Details
        driver.find_element(By.XPATH, "//button[text()= 'Next']").click()


        driver.find_element(By.XPATH, "//b[text()= '+ New Regulation']").click()


        countryselect = driver.find_element(By.XPATH, "//select[@name= 'country']")
        Select(countryselect).select_by_index(1)


        selectRegulation = driver.find_element(By.XPATH, "//select[@name= 'regulation']")
        Select(selectRegulation).select_by_index(1)


        driver.find_element(By.XPATH, "//input[@placeholder= 'Enter weight (gm)']").send_keys(weight)

        driver.find_element(By.XPATH, "//input[@name='strength (mg)']").send_keys(strength)


        driver.find_element(By.XPATH, "//button[text()= 'Accept']").click()


        # Next to Custom Details
        driver.find_element(By.XPATH, "//button[text()= 'Next']").click()


        # Submit Details
        driver.find_element(By.XPATH, "//button[text()= 'Submit']").click()

if __name__ == "__main__":
    main()
