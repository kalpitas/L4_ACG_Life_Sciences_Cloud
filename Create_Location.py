import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def main():

    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.implicitly_wait(60)

    workbook = load_workbook("ACG_Common_Workbook.xlsx")
    sheet = workbook["partner"]
    url_login_cred = workbook["URL_Login_cred_Tenant"]
    
    start_row = 2

    time.sleep(3)
    driver.get("https://proud-mud-040601f00.4.azurestaticapps.net/")

    driver.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

    usern = url_login_cred.cell(row=2, column=2).value
    # if type_env == 'ACG':
    driver.find_element(By.NAME, 'userName').send_keys(str(usern))

    passw = url_login_cred.cell(row=2, column=3).value
    driver.find_element(By.NAME, 'password').send_keys(str(passw))

    login = driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[4]/div/div[2]/div/button")
    login.click()


    driver.find_element(By.XPATH,"/html/body/div[1]/div/div/div[2]/div[1]/div/div[2]/h2/button/span").click()
    driver.find_element(By.XPATH,"/html/body/div[1]/div/div/div[2]/div[1]/div/div[2]/div/div/a[3]").click()


    # Loop through all the rows in the Excel sheet
    for i in range(start_row, sheet.max_row + 1):  # Adjusting for 0-indexing
        # Click 'New Location' button at the beginning of each loop

        timeout = 10
        try:
            new_role = EC.element_to_be_clickable((By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b"))
            WebDriverWait(driver,timeout).until(new_role)
        except TimeoutException:
            print("Create role page: Timed out waiting for page to load")
            
        time.sleep(5)
        
        driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b").click()


        location_name = sheet.cell(row = i, column = 1).value
        location_id = sheet.cell(row = i, column = 2).value
        state = sheet.cell(row = i, column = 3).value
        city = sheet.cell(row = i, column = 4).value
        address = sheet.cell(row = i, column = 5).value
        postal_code = sheet.cell(row = i, column = 6).value
        contact_person = sheet.cell(row = i, column = 7).value
        email_id = sheet.cell(row = i, column = 8).value
        phone_number = sheet.cell(row = i, column = 9).value
        website = sheet.cell(row = i, column = 10).value
        entity = sheet.cell(row = i, column = 11).value
        bus_entity = sheet.cell(row = i, column = 12).value
        loc_identifier_ty = sheet.cell(row = i, column = 13).value

        # Fill in the location details
        locname = driver.find_element(By.XPATH, "//input[@placeholder='Enter location name']")
        locname.send_keys(location_name)

        loc_type = Select(driver.find_element(By.XPATH, "//select[@name='locationType']"))
        loc_type.select_by_value("Physical Site")

        entity1 = Select(driver.find_element(By.XPATH, "//select[@name='entity']"))
        entity1.select_by_visible_text(str(entity))

        business_entity = Select(driver.find_element(By.XPATH, "//select[@name='locName']"))
        business_entity.select_by_visible_text(str(bus_entity))
        
        time.sleep(2)
        
        element = driver.find_element(By.XPATH, "//select[@name='locationIdType']")
        driver.execute_script("arguments[0].scrollIntoView(true);", element)

        id_type = Select(driver.find_element(By.XPATH, "//select[@name='locationIdType']"))
        id_type.select_by_visible_text(str(loc_identifier_ty))

        identifier = driver.find_element(By.XPATH, "//input[@name='locationId']")
        identifier.send_keys(location_id)
        
        time.sleep(2)

        next1 = driver.find_element(By.XPATH, "//button[text()='Next']")
        next1.click()


        # Fill in the address details
        country = Select(driver.find_element(By.XPATH, "//select[@name='country']"))
        country.select_by_value("India")


        state_element = driver.find_element(By.XPATH, "//input[@name='state']")
        state_element.send_keys(state)


        city_element = driver.find_element(By.XPATH, "//input[@name='city']")
        city_element.send_keys(city)


        address_element = driver.find_element(By.XPATH, "//input[@name='address']")
        address_element.send_keys(address)


        pcode = driver.find_element(By.XPATH, "//input[@name='postalCode']")
        pcode.send_keys(postal_code)


        # Scroll into view and fill in contact details
        element = driver.find_element(By.XPATH, "//input[@name='contactPersonName']")
        driver.execute_script("arguments[0].scrollIntoView(true);", element)


        conper = driver.find_element(By.XPATH, "//input[@placeholder='Enter name']")
        conper.send_keys(contact_person)


        email_element = driver.find_element(By.XPATH, "//input[@placeholder='Enter email']")
        email_element.send_keys(email_id)


        phno = driver.find_element(By.XPATH, "//input[@placeholder='Enter phone number']")
        phno.send_keys(phone_number)


        website_element = driver.find_element(By.XPATH, "//input[@placeholder='Enter website']")
        website_element.send_keys(website)


        # Submit the form
        submit = driver.find_element(By.XPATH, "//button[text()='Submit']")
        submit.click()

        time.sleep(3)



if __name__ == "__main__":
    main()