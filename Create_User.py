from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

    
driver = webdriver.Chrome()
driver.maximize_window()
#driver.implicitly_wait(60)

        
wb = load_workbook("ACG_Common_Workbook.xlsx")
typeev = wb["URL_Login_cred_Tenant"]
type_env = typeev.cell(row = 2, column = 4).value



if type_env == "ACG":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_ACG"]

if type_env == "Tenant":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_Tenant"]


url = url_login_cred.cell(row = 2, column = 1).value
print(url)
driver.get(str(url))

#if type_env == 'ACG':
driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()


usern = url_login_cred.cell(row = 2, column = 2).value
#if type_env == 'ACG':
driver.find_element(By.NAME,'userName').send_keys(str(usern))

#if type_env == 'Tenant':
#    driver.find_element(By.NAME,'userName').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[2]/div/button").click()

user = wb['Users']
start_user = 2
last_user = (user.max_row) +1

#driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[1]/div/div/button[2]").click()
#driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[1]/div/div/button[2]/div").click()
if type_env == 'ACG':
    driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[2]/h2/button/span').click()
    driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[2]/div/div/a[1]').click()

if type_env == 'Tenant':
    driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[1]/h2/button/span').click()
    driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[1]/div/div/a[1]').click()
    

for i in range(start_user, last_user):

    timeout = 10
    try:
        new_user_nxt = EC.element_to_be_clickable((By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div[1]/button[2]"))
        WebDriverWait(driver,timeout).until(new_user_nxt)
    except TimeoutException:
        print("Create user page_nxt: Timed out waiting for page to load")
    
    
    name = user.cell(row = i, column = 1).value
    driver.find_element(By.NAME,"userName").send_keys(str(name))
    
    email = user.cell(row = i, column = 2).value
    driver.find_element(By.NAME,"emailAddress").send_keys(str(email))
    
    mob = user.cell(row = i, column = 3).value
    driver.find_element(By.NAME,"phoneNumber").send_keys(str(mob))
    
    # select by visible text
    role = user.cell(row = i, column = 4).value
    print(role)
    select = Select(driver.find_element(By.NAME,"roleId"))
    select.select_by_visible_text(str(role))
    
    if type_env == 'Tenant':
        location = user.cell(row = i, column = 5).value
        print(location)
        select = Select(driver.find_element(By.NAME,"locationId"))
        select.select_by_visible_text(str(location))
        
    '''
    
    driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[3]/div[1]/button[2]").click()
    
    timeout = 10
    try:
        new_user_sub = EC.element_to_be_clickable((By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]"))
        WebDriverWait(driver,timeout).until(new_user_sub)
    except TimeoutException:
        print("Create user page_sub: Timed out waiting for page to load")
        
    driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]").click()
    
    print("row", i, "got completed successfully")
    
    timeout = 10
    try:
        new_user = EC.element_to_be_clickable((By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b"))
        WebDriverWait(driver,timeout).until(new_user)
    except TimeoutException:
        print("Create new user: Timed out waiting for page to load")
        
    driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b").click()
    
    '''
        
    #click on next button
    driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[3]/div[1]/button[2]").click()
    print("clicked on next button")

    timeout = 2
    try:
        #click on submit
        new_user_sub = EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]"))
        WebDriverWait(driver, timeout).until(new_user_sub)
        driver.find_element(By.XPATH, "//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]").click()
        print("clicked on submit button")
        try:
            # click on new user
            new_user = EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b"))
            WebDriverWait(driver, timeout).until(new_user)
            #successfull message
            message = driver.find_element(By.XPATH,"/html/body/div[2]/div[4]/div/div/div/div/div[2]").text
            print(message)
        except TimeoutException:
            print("Nested Final exception")
            driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[3]/div[1]/button[2]").click()
            driver.find_element(By.XPATH, "//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]").click()
            message = driver.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/div/div/div[2]").text
            print(message)
            driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div[1]/button[1]").click()
    except TimeoutException:
        print("Final exception")
        driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[3]/div[1]/button[2]").click()
        message = driver.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/div/div/div[2]").text
        print(message)
        driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div[1]/button[1]").click()

    user.cell(row=i, column=6).value = message
    wb.save("ACG_Common_Workbook.xlsx")

    time.sleep(5)

    driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b").click()
    print("clicked on new user button")
