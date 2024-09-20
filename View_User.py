from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

    
driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(60)

        
wb = load_workbook("ACG_Common_Workbook.xlsx")
typeev = wb["URL_Login_cred_Tenant"]
type_env = typeev.cell(row = 2, column = 4).value



if type_env == "ACG":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_ACG"]

if type_env == "Tenant":
    #fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_Tenant"]


url = url_login_cred.cell(row = 2, column = 1).value
print(url)
driver.get(str(url))

#if type_env == 'ACG':
driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()


usern = url_login_cred.cell(row = 2, column = 2).value
#if type_env == 'ACG':
driver.find_element(By.NAME,'userName').send_keys(str(usern))

#if type_env == 'Tenant':
#    driver.find_element(By.NAME,'userName').send_keys(str(usern))

passw = url_login_cred.cell(row = 2, column = 3).value
driver.find_element(By.NAME,'password').send_keys(str(passw))

driver.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[2]/div/button").click()

user = wb['Users']
start_user = 2
last_user = (user.max_row) +1

#driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[1]/div/div/button[2]").click()
#driver.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[1]/div/div/button[2]/div").click()
if type_env == 'ACG':
    driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[2]/h2/button/span').click()
    driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[2]/div/div/a[2]').click()

if type_env == 'Tenant':
    driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[1]/h2/button/span').click()
    driver.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[1]/div/div/a[2]').click()
    

for i in range(start_user, last_user):

    #waiting for new user 
    timeout = 10
    try:
        new_user = EC.element_to_be_clickable((By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b"))
        WebDriverWait(driver,timeout).until(new_user)
    except TimeoutException:
        print("View user page: Timed out waiting for page to load")
    

    name = user.cell(row = i, column = 1).value
    search = driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/input")
    search.click()
    search.send_keys(str(name))

    #email = user.cell(row = i, column = 2).value
    #driver.find_element(By.NAME,"emailAddress").send_keys(str(email))
     

    timeout = 5
    try:
        #click on submit
        user_view = EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/div[2]/div[2]/div/div/div[3]/div/table/tbody/tr/td[6]/button"))
        WebDriverWait(driver, timeout).until(user_view)
        usern = driver.find_element(By.XPATH, "//*[@id='root']/div/div/div[2]/div[2]/div/div/div[3]/div/table/tbody/tr/td[2]").text
        user.cell(row = i, column = 8).value = usern

        emailid= driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[3]/div/table/tbody/tr/td[3]").text
        user.cell(row = i, column = 9).value = emailid
        
        rolen = driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[3]/div/table/tbody/tr/td[4]").text
        user.cell(row = i, column = 10).value = rolen
        
        statuss = driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[3]/div/table/tbody/tr/td[5]").text
        user.cell(row = i, column = 11).value = statuss
        
    except TimeoutException:
        user.cell(row = i, column = 8).value = "User not created"

    wb.save("ACG_Common_Workbook.xlsx")
    
    time.sleep(5)

    driver.refresh()

    #driver.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b").click()
    #print("clicked on new user button")
